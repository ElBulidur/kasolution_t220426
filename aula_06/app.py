# CSV
import csv
import openpyxl
# from openpyxl import load_workbook

alunos_csv = [
    ["Julio", 100]
    , ["Ana", 80]
    , ["Roberta", 78]
    , ["Isaque", 95]
    , ["Leonardo", 10]
]

# criando
# with open("alunos.csv", "w", newline='') as file:
#     gravar = csv.writer(file, delimiter=";")
#     gravar.writerow(["Aluno", "Notas"])
#     gravar.writerows(alunos_csv)
    
arquivo = "alunos.csv"

# with open("alunos.csv", "r") as file:
#     ler = csv.reader(file)
#     for linha in ler:
#         print(linha)


# EXCEL

# arquivo = "alunos.xlsx"

# wb = openpyxl.load_workbook(arquivo)
# # ws = wb.active
# ws = wb['Prazos']

# # print(wb.sheetnames)
# for linha in ws.iter_rows(min_row=1, max_row=10, values_only=True):
#     print(linha)



#  CRIANDO ARQUIVO EXCEL
# arquivo_novo = "alunos_novo.xlsx"

# wb = openpyxl.Workbook()

# ws = wb.active
# ws.title = "lista de alunos"

# ws['A1'] = "ALUNO"
# ws['B1'] = "NOTA"

# ws['A2'] = "Julio Pereira"
# ws['B2'] = "100"

# ws['A3'] = "Maria"
# ws['B3'] = "10"

# wb.save(arquivo_novo)


# PANDAS
# NUMPY
# SCIPY
# MATPLOTLIB
import pandas as pd

excel = pd.read_excel("alunos.xlsx", sheet_name="Prazos")

# print(excel.head())


df = pd.DataFrame(
    {
        "Name": [
            "Braund, Mr. Owen Harris",
            "Allen, Mr. William Henry",
            "Bonnell, Miss Elizabeth",
        ],
        "Age": [22, 35, 58],
        "Sex": ["male", "male", "female"],
    }
)

# df.to_excel("exemplo.xlsx")

# print(df.head())

